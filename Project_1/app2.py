import openpyxl as xl
from openpyxl.chart import BarChart, Reference


wb = xl.load_workbook('transactions.xlsx')  # first creating this function to load a workbook
sheet = wb['Sheet1'] #wb creates an object and now we are going to access the sheet.
"""cell = sheet['a1'] # to access the a1 cell
cell = sheet.cell(1,1) # this is doing the same thing but it is hard to access if we have multiple cells
"""
#print(cell.value) # THis will print transaction_id
#print(sheet.max_row) # 4


for row in range (2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price
values = Reference (sheet,
           min_row=2,
           max_row=sheet.max_row,
           min_col=4,
           max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('transaction3.xlsx') # This will save the changes and create a new file called transaction2

